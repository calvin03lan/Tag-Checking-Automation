"""
Excel config adapter.

This module is the single place responsible for translating a structured
Excel input file into application data models.

To support different Excel formats later, modify the schema in this module.
"""
from dataclasses import dataclass
from pathlib import Path
from typing import List, Tuple

import openpyxl

from models.config import LANG_COMPAT_MAP, LANG_MAP
from models.session import KeywordItem, UrlItem
from utils.new_template_adapter import NewTemplateConfigAdapter


@dataclass(frozen=True)
class ExcelConfigSchema:
    """Defines sheet names and column indices for Excel parsing."""

    urls_sheet: str = "URLs"
    keywords_sheet: str = "Keywords"

    # URLs row: num | lang | url
    urls_num_col: int = 0
    urls_lang_col: int = 1
    urls_url_col: int = 2

    # Keywords row: num | lang | text | button_name
    kw_num_col: int = 0
    kw_lang_col: int = 1
    kw_text_col: int = 2
    kw_button_col: int = 3


class ExcelConfigAdapter:
    """
    Converts Excel file content into URL/Keyword model objects.
    """

    def __init__(self, schema: ExcelConfigSchema | None = None) -> None:
        self.schema = schema or ExcelConfigSchema()
        self._new_template_adapter = NewTemplateConfigAdapter()

    def load_to_models(self, filepath: str) -> Tuple[List[UrlItem], List[KeywordItem]]:
        path = Path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"Config file not found: {filepath}")

        wb = openpyxl.load_workbook(filepath, data_only=True)
        try:
            if self._is_exact_legacy_workbook(wb.sheetnames):
                url_items = self._parse_urls_sheet(wb[self.schema.urls_sheet])
                keyword_items = self._parse_keywords_sheet(wb[self.schema.keywords_sheet])
                return url_items, keyword_items

            # Not the exact legacy 2-sheet structure -> parse as new template.
            return self._new_template_adapter.load_to_models(wb.active)
        finally:
            wb.close()

    def _is_exact_legacy_workbook(self, sheet_names: List[str]) -> bool:
        expected = {self.schema.urls_sheet, self.schema.keywords_sheet}
        return len(sheet_names) == 2 and set(sheet_names) == expected

    def load_legacy_keywords(self, filepath: str) -> List[str]:
        """
        Legacy helper: read column A of first worksheet (skip row 1).
        """
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        try:
            ws = wb.active
            keywords: List[str] = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    continue
                value = row[0] if row else None
                if value is not None:
                    kw = str(value).strip()
                    if kw:
                        keywords.append(kw)
            return keywords
        finally:
            wb.close()

    def _parse_urls_sheet(self, ws) -> List[UrlItem]:
        items: List[UrlItem] = []
        s = self.schema
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                continue
            if not row or len(row) <= s.urls_url_col or row[s.urls_url_col] is None:
                continue
            num_cell = row[s.urls_num_col] if len(row) > s.urls_num_col else None
            lang_label = row[s.urls_lang_col] if len(row) > s.urls_lang_col else None
            url = row[s.urls_url_col]
            url_str = str(url).strip()
            if not url_str:
                continue
            num = int(num_cell) if num_cell is not None else 1
            lang_code = _normalize_lang(lang_label)
            items.append(UrlItem(url=url_str, lang=lang_code, num=num))
        return items

    def _parse_keywords_sheet(self, ws) -> List[KeywordItem]:
        items: List[KeywordItem] = []
        s = self.schema
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                continue
            if not row or len(row) <= s.kw_text_col or row[s.kw_text_col] is None:
                continue
            num_cell = row[s.kw_num_col] if len(row) > s.kw_num_col else None
            lang_cell = row[s.kw_lang_col] if len(row) > s.kw_lang_col else None
            text_cell = row[s.kw_text_col]
            button_cell = row[s.kw_button_col] if len(row) > s.kw_button_col else None

            num = int(num_cell) if num_cell is not None else 0
            lang_label = str(lang_cell).strip() if lang_cell is not None else ""
            text = str(text_cell).strip()
            button_name = str(button_cell).strip() if button_cell is not None else None
            if not text:
                continue
            lang_code = _normalize_lang(lang_label)
            items.append(
                KeywordItem(num=num, lang=lang_code, text=text, button_name=button_name)
            )
        return items


def load_excel_to_models(
    filepath: str, schema: ExcelConfigSchema | None = None
) -> Tuple[List[UrlItem], List[KeywordItem]]:
    return ExcelConfigAdapter(schema=schema).load_to_models(filepath)


def load_legacy_keywords(filepath: str) -> List[str]:
    return ExcelConfigAdapter().load_legacy_keywords(filepath)


def _normalize_lang(raw_lang) -> str:
    """
    Normalize language labels/aliases to one of: tc/sc/en.
    """
    raw = str(raw_lang).strip() if raw_lang is not None else ""
    for cand in _lang_lookup_candidates(raw):
        if cand in LANG_COMPAT_MAP:
            return LANG_COMPAT_MAP[cand]
    # Backward compatibility with existing friendly labels in models.config.LANG_MAP.
    return LANG_MAP.get(raw, "en")


def _lang_lookup_candidates(raw: str) -> List[str]:
    """
    Generate normalized variants to improve alias matching robustness.
    """
    if not raw:
        return [""]
    base = raw.strip().lower()
    variants = [
        base,
        base.replace("_", "-"),
        base.replace("-", "_"),
        base.replace(" ", "-"),
        base.replace(" ", "_"),
        base.replace(" ", ""),
    ]
    # De-duplicate while preserving order.
    return list(dict.fromkeys(variants))
