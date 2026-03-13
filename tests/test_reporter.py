"""Unit tests for core/reporter.py"""
import pytest
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from PIL import Image
from openpyxl import load_workbook

from core.reporter import ReportWriter
from models.session import ReportEntry


def _make_screenshot(path: str) -> str:
    Image.new("RGB", (320, 180), (100, 149, 237)).save(path)
    return path


def _make_input_template(path: str) -> str:
    wb = load_workbook(path) if Path(path).exists() else None
    if wb:
        wb.close()
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Matching rule requirements"
    ws["B3"] = "DoubleClick"
    ws["C3"] = "Gtag"
    ws["D3"] = "Meta"
    ws["E3"] = "The Trade Desk"
    ws["F3"] = "Taboola"
    ws["G3"] = "Applier"
    ws["B4"] = "Required"
    ws["C4"] = "Required"
    ws["D4"] = "Not Required"
    ws["K4"] = "/zh-hk"
    ws["L4"] = "/promo/a"
    ws["B5"] = "Required"
    ws["K5"] = "/en-hk"
    ws["L5"] = "/promo/b"
    ws["A190"] = "   "
    ws["AC190"] = "helper"
    wb.save(path)
    wb.close()
    return path


def _read_first_image_size_px(xlsx_path: str) -> tuple[int, int]:
    xdr_ns = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    a_ns = "http://schemas.openxmlformats.org/drawingml/2006/main"

    with zipfile.ZipFile(xlsx_path, "r") as zf:
        drawing_files = sorted(
            name
            for name in zf.namelist()
            if name.startswith("xl/drawings/drawing") and name.endswith(".xml")
        )
        if not drawing_files:
            raise AssertionError("No drawing xml found in workbook")

        data = zf.read(drawing_files[0])
        root = ET.fromstring(data)

    ext = root.find(f".//{{{xdr_ns}}}oneCellAnchor/{{{xdr_ns}}}ext")
    if ext is None:
        ext = root.find(f".//{{{a_ns}}}ext")
    if ext is None:
        raise AssertionError("No image size ext found in drawing xml")

    width_px = int(round(int(ext.attrib["cx"]) / 9525.0))
    height_px = int(round(int(ext.attrib["cy"]) / 9525.0))
    return width_px, height_px


def _make_entry(
    url_index=1, url="https://example.com", url_lang="en",
    kw_num=1, kw_text="gtm_click", kw_lang="en", kw_button=None,
    result="PASS", tested_at="2024-01-01 10:00:00", screenshot_path=None,
    tag_vendor="other", source_row=0,
):
    return ReportEntry(
        url_index=url_index, url=url, url_lang=url_lang,
        kw_num=kw_num, kw_text=kw_text, kw_lang=kw_lang, kw_button=kw_button,
        result=result, tested_at=tested_at, screenshot_path=screenshot_path,
        tag_vendor=tag_vendor, source_row=source_row,
    )


def _basic_entries():
    return [
        _make_entry(
            url_index=1, url="https://example.com", url_lang="en",
            kw_num=1, kw_text="gtm_click", kw_lang="en", kw_button="btn-a",
            result="PASS", tested_at="2024-01-01 10:00:00",
        ),
        _make_entry(
            url_index=1, url="https://example.com", url_lang="en",
            kw_num=2, kw_text="pageview", kw_lang="tc", kw_button=None,
            result="FAILED", tested_at="2024-01-01 10:00:05",
        ),
    ]


class TestReportWriter:
    def test_file_is_created(self, tmp_path):
        out = str(tmp_path / "report.xlsx")
        ReportWriter().write(_basic_entries(), out)
        assert Path(out).exists()

    def test_returns_absolute_path(self, tmp_path):
        out    = str(tmp_path / "report.xlsx")
        result = ReportWriter().write(_basic_entries(), out)
        assert Path(result).is_absolute()

    def test_header_row_values(self, tmp_path):
        out = str(tmp_path / "report.xlsx")
        ReportWriter().write(_basic_entries(), out)
        ws = load_workbook(out).active
        headers = [ws.cell(1, c).value for c in range(1, 11)]
        assert headers == [
            "URL #", "URL", "URL Lang",
            "KW #", "Keyword", "KW Lang", "ID",
            "Result", "Tested At", "Screenshot",
        ]

    def test_data_rows_written(self, tmp_path):
        out = str(tmp_path / "report.xlsx")
        ReportWriter().write(_basic_entries(), out)
        ws = load_workbook(out).active
        # Row 2: first keyword entry
        assert ws["A2"].value == 1           # URL #
        assert ws["B2"].value == "https://example.com"
        assert ws["E2"].value == "gtm_click" # Keyword
        assert ws["G2"].value == "btn-a"     # ID
        assert ws["H2"].value == "PASS"      # Result
        # Row 3: second keyword entry
        assert ws["E3"].value == "pageview"
        assert ws["G3"].value == "—"         # no button → "—"
        assert ws["H3"].value == "FAILED"

    def test_url_lang_label_translated(self, tmp_path):
        out = str(tmp_path / "report.xlsx")
        ReportWriter().write(_basic_entries(), out)
        ws = load_workbook(out).active
        assert ws["C2"].value == "English"

    def test_kw_lang_label_translated(self, tmp_path):
        out = str(tmp_path / "report.xlsx")
        ReportWriter().write(_basic_entries(), out)
        ws = load_workbook(out).active
        assert ws["F2"].value == "English"
        assert ws["F3"].value == "Traditional Chinese"

    def test_empty_entries_produces_header_only(self, tmp_path):
        out = str(tmp_path / "report.xlsx")
        ReportWriter().write([], out)
        ws = load_workbook(out).active
        assert ws["A2"].value is None

    def test_with_screenshot_no_exception(self, tmp_path):
        shot = _make_screenshot(str(tmp_path / "shot.png"))
        entries = [_make_entry(result="PASS", screenshot_path=shot)]
        out = str(tmp_path / "report.xlsx")
        ReportWriter().write(entries, out)
        assert Path(out).exists()

    def test_missing_screenshot_writes_na(self, tmp_path):
        entries = [_make_entry(
            result="PASS", screenshot_path="/non/existent/shot.png"
        )]
        out = str(tmp_path / "report.xlsx")
        ReportWriter().write(entries, out)
        ws = load_workbook(out).active
        assert ws["J2"].value == "N/A"

    def test_new_template_sheet_is_created_with_12_columns(self, tmp_path):
        out = str(tmp_path / "report.xlsx")
        ReportWriter().write(_basic_entries(), out)
        wb = load_workbook(out)
        assert "New Template Report" in wb.sheetnames
        ws = wb["New Template Report"]
        headers = [ws.cell(1, c).value for c in range(1, 13)]
        assert headers == [
            "DoubleClick", "Evidence",
            "Gtag", "Evidence",
            "Meta", "Evidence",
            "The Trade Desk", "Evidence",
            "Taboola", "Evidence",
            "Applier", "Evidence",
        ]

    def test_new_template_sheet_respects_source_row_order(self, tmp_path):
        out = str(tmp_path / "report.xlsx")
        entries = [
            _make_entry(
                kw_text="dc_b", result="FAILED", tag_vendor="doubleclick", source_row=8
            ),
            _make_entry(
                kw_text="dc_a", result="PASS", tag_vendor="doubleclick", source_row=4
            ),
            _make_entry(
                kw_text="gtag_a", result="PASS", tag_vendor="gtag", source_row=4
            ),
        ]
        ReportWriter().write(entries, out)
        wb = load_workbook(out)
        ws = wb["New Template Report"]
        assert ws["A2"].value == "PASS"
        assert ws["C2"].value == "PASS"
        assert ws["A3"].value == "FAILED"

    def test_export_into_input_copy_updates_summary_and_writes_evidence_at_ac(self, tmp_path):
        input_path = _make_input_template(str(tmp_path / "input.xlsx"))
        out = str(tmp_path / "output.xlsx")
        entries = [
            _make_entry(result="PASS", tag_vendor="doubleclick", source_row=4),
            _make_entry(result="FAILED", tag_vendor="gtag", source_row=4),
            _make_entry(result="PASS", tag_vendor="doubleclick", source_row=5),
        ]

        saved = ReportWriter().write(entries, out)  # baseline API still works
        assert Path(saved).exists()

        saved2 = ReportWriter().write(entries, out)
        assert Path(saved2).exists()

        from core.excel_exporter import ExcelReportExporter
        saved3 = ExcelReportExporter().export_into_input_copy(entries, input_path, out)
        assert Path(saved3).exists()

        wb = load_workbook(out)
        try:
            assert wb.sheetnames[0] == "Summary"
            assert wb.sheetnames[1] == "Test Result"
            ws = wb["Matching rule requirements"]
            assert ws["B4"].value == "Required"
            assert ws["C4"].value == "Required"
            assert ws["D4"].value == "Not Required"
            assert ws["B5"].value == "Required"

            ws_overview = wb["Summary"]
            assert ws_overview["A1"].value == "Summary"
            assert ws_overview["A2"].value == "Latest Update"
            assert ws_overview["A3"].value == "Status"
            assert re.match(r"^\d{4}/\d{2}/\d{2}$", str(ws_overview["B2"].value))
            assert ws_overview["B3"].value in {None, ""}

            ws_result = wb["Test Result"]
            assert ws_result["B4"].value == "PASS"
            assert ws_result["C4"].value == "FAILED"
            assert ws_result["D4"].value == "Not Required"
            assert ws_result["B5"].value == "PASS"

            headers = [ws_result.cell(3, c).value for c in range(29, 41)]
            assert headers == [
                "DoubleClick", "Evidence",
                "Gtag", "Evidence",
                "Meta", "Evidence",
                "The Trade Desk", "Evidence",
                "Taboola", "Evidence",
                "Applier", "Evidence",
            ]
            assert ws_result["AC4"].value == "PASS"
            assert ws_result["AE4"].value == "FAILED"
            assert ws_result["AC5"].value == "PASS"
            assert ws_result["B4"].fill.fgColor.rgb == "FF92D050"
            assert ws_result["B4"].font.bold is True
            assert ws_result["B4"].alignment.horizontal == "center"
            assert ws_result["B4"].alignment.vertical == "center"
            assert ws_result["C4"].fill.fgColor.rgb == "FFFF0000"
            assert ws_result["C4"].font.bold is True
            assert ws_result["C4"].alignment.horizontal == "center"
            assert ws_result["C4"].alignment.vertical == "center"

            assert ws_result["AC4"].fill.fgColor.rgb == "FF92D050"
            assert ws_result["AC4"].font.bold is True
            assert ws_result["AC4"].alignment.horizontal == "center"
            assert ws_result["AC4"].alignment.vertical == "center"
            assert ws_result["AE4"].fill.fgColor.rgb == "FFFF0000"
            assert ws_result["AE4"].font.bold is True
            assert ws_result["AE4"].alignment.horizontal == "center"
            assert ws_result["AE4"].alignment.vertical == "center"
            assert ws_result["AC3"].alignment.horizontal == "center"
            assert ws_result["AC3"].alignment.vertical == "bottom"

            assert ws_result.row_dimensions[4].height == 90
            assert ws_result.row_dimensions[5].height == 90
            assert ws_result["A6"].value is None

            assert "Evidence" not in wb.sheetnames
            assert "Tag QA Report" not in wb.sheetnames
            assert ws_result.column_dimensions["AD"].width == 51
            assert ws_result.column_dimensions["AF"].width == 51
            assert ws_result.column_dimensions["AH"].width == 51
            assert ws_result.column_dimensions["AJ"].width == 51
            assert ws_result.column_dimensions["AL"].width == 51
            assert ws_result.column_dimensions["AN"].width == 51
        finally:
            wb.close()

    def test_report_writer_write_into_input_copy(self, tmp_path):
        input_path = _make_input_template(str(tmp_path / "input.xlsx"))
        out = str(tmp_path / "output.xlsx")
        entries = [
            _make_entry(result="PASS", tag_vendor="doubleclick", source_row=4),
        ]

        saved = ReportWriter().write_into_input_copy(
            entries,
            input_excel_path=input_path,
            output_path=out,
        )
        assert Path(saved).exists()

        wb = load_workbook(out)
        try:
            assert wb.sheetnames[0] == "Summary"
            assert wb.sheetnames[1] == "Test Result"
        finally:
            wb.close()

    def test_export_into_input_copy_inserts_two_rows_for_all_language(self, tmp_path):
        input_path = _make_input_template(str(tmp_path / "input.xlsx"))
        wb = load_workbook(input_path)
        ws = wb["Matching rule requirements"]
        ws["K4"] = "(All)"
        wb.save(input_path)
        wb.close()

        out = str(tmp_path / "output.xlsx")
        entries = [
            _make_entry(result="PASS", tag_vendor="doubleclick", source_row=4, url_lang="tc"),
            _make_entry(result="FAILED", tag_vendor="doubleclick", source_row=4, url_lang="sc"),
            _make_entry(result="PASS", tag_vendor="doubleclick", source_row=4, url_lang="en"),
            _make_entry(result="PASS", tag_vendor="doubleclick", source_row=5, url_lang="en"),
        ]

        from core.excel_exporter import ExcelReportExporter
        saved = ExcelReportExporter().export_into_input_copy(entries, input_path, out)
        assert Path(saved).exists()

        wb = load_workbook(out)
        try:
            ws_result = wb["Test Result"]
            assert ws_result["K4"].value == "(All)"
            assert ws_result["K5"].value is None
            assert ws_result["K6"].value is None
            assert ws_result["K7"].value == "/en-hk"

            assert ws_result["AC4"].value == "PASS"
            assert ws_result["AC5"].value == "FAILED"
            assert ws_result["AC6"].value == "PASS"
            assert ws_result["AC7"].value == "PASS"

            assert ws_result.row_dimensions[4].height == 90
            assert ws_result.row_dimensions[5].height == 90
            assert ws_result.row_dimensions[6].height == 90
            assert ws_result.row_dimensions[7].height == 90

            assert ws_result["B4"].value == "FAILED"
            assert ws_result["B7"].value == "PASS"
            assert ws_result["A8"].value is None
        finally:
            wb.close()

    def test_export_into_input_copy_image_size_matches_cell_size(self, tmp_path):
        input_path = _make_input_template(str(tmp_path / "input.xlsx"))
        out = str(tmp_path / "output.xlsx")
        shot = _make_screenshot(str(tmp_path / "shot.png"))
        entries = [
            _make_entry(
                result="PASS",
                tag_vendor="doubleclick",
                source_row=4,
                screenshot_path=shot,
            ),
        ]

        from core.excel_exporter import ExcelReportExporter
        saved = ExcelReportExporter().export_into_input_copy(entries, input_path, out)
        assert Path(saved).exists()
        width_px, height_px = _read_first_image_size_px(out)
        assert width_px == 362
        assert height_px == 120

    def test_export_into_input_copy_summary_sheet_marks_all_pass(self, tmp_path):
        input_path = _make_input_template(str(tmp_path / "input.xlsx"))
        out = str(tmp_path / "output.xlsx")
        entries = [
            _make_entry(result="PASS", tag_vendor="doubleclick", source_row=4),
            _make_entry(result="PASS", tag_vendor="gtag", source_row=4),
        ]

        from core.excel_exporter import ExcelReportExporter
        saved = ExcelReportExporter().export_into_input_copy(entries, input_path, out)
        assert Path(saved).exists()

        wb = load_workbook(out)
        try:
            ws_overview = wb["Summary"]
            assert ws_overview["A1"].value == "Summary"
            assert ws_overview["A2"].value == "Latest Update"
            assert ws_overview["A3"].value == "Status"
            assert re.match(r"^\d{4}/\d{2}/\d{2}$", str(ws_overview["B2"].value))
            assert "All PASS" in str(ws_overview["B3"].value)
        finally:
            wb.close()
