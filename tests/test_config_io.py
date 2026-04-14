"""Unit tests for utils/config_io.py"""
import pytest
import openpyxl
from pathlib import Path

from utils.config_io import convert_excel_to_json_data, load_excel_config


# ── Helpers ──────────────────────────────────────────────────────────

def _make_two_sheet_excel(tmp_path, url_rows, kw_rows):
    """Build a minimal two-sheet Excel file for load_excel_config."""
    wb = openpyxl.Workbook()
    ws_urls = wb.active
    ws_urls.title = "URLs"
    for row in url_rows:
        ws_urls.append(row)

    ws_kw = wb.create_sheet("Keywords")
    for row in kw_rows:
        ws_kw.append(row)

    path = str(tmp_path / "config.xlsx")
    wb.save(path)
    return path


def _make_new_template_excel(tmp_path, rows):
    """Build a minimal new-template one-sheet Excel for auto-fallback parsing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matching rule requirements"

    ws["B3"] = "DoubleClick\n(Fill in by MKG)"
    ws["C3"] = "Gtag\n(Fill in by MKG)"
    ws["D3"] = "Meta\n(Fill in by MKG)"
    ws["E3"] = "The Trade Desk\n(Fill in by MKG)"
    ws["F3"] = "Taboola\n(Fill in by MKG)"
    ws["G3"] = "Appier\n(Fill in by MKG)"
    ws["H3"] = "Page or event?\n(Fill in by MKG)"
    ws["K3"] = "Language\n[Read Note on this cell]"
    ws["L3"] = "URL path name\n[without domain and language code]"
    ws["P3"] = "Button ID - for event only"
    ws["Q3"] = "type(Group Tag String)"
    ws["R3"] = "cat (Activity Tag String)"
    ws["S3"] = "Gtag Event snippet (Conversion ID / Conversion label)"
    ws["T3"] = "Master Pixel ID"
    ws["U3"] = "EV value"
    ws["V3"] = "Account ID"
    ws["W3"] = "CT value"
    ws["X3"] = "Account ID"
    ws["Y3"] = "EN value"
    ws["Z3"] = "type"
    ws["AA3"] = "action_id"
    ws["AB3"] = "track_id"

    for idx, item in enumerate(rows, start=4):
        ws[f"B{idx}"] = item.get("B")
        ws[f"C{idx}"] = item.get("C")
        ws[f"D{idx}"] = item.get("D")
        ws[f"E{idx}"] = item.get("E")
        ws[f"F{idx}"] = item.get("F")
        ws[f"G{idx}"] = item.get("G")
        ws[f"H{idx}"] = item.get("H")
        ws[f"K{idx}"] = item.get("K")
        ws[f"L{idx}"] = item.get("L")
        ws[f"P{idx}"] = item.get("P")
        ws[f"Q{idx}"] = item.get("Q")
        ws[f"R{idx}"] = item.get("R")
        ws[f"S{idx}"] = item.get("S")
        ws[f"T{idx}"] = item.get("T")
        ws[f"U{idx}"] = item.get("U")
        ws[f"V{idx}"] = item.get("V")
        ws[f"W{idx}"] = item.get("W")
        ws[f"X{idx}"] = item.get("X")
        ws[f"Y{idx}"] = item.get("Y")
        ws[f"Z{idx}"] = item.get("Z")
        ws[f"AA{idx}"] = item.get("AA")
        ws[f"AB{idx}"] = item.get("AB")

    path = str(tmp_path / "new_template.xlsx")
    wb.save(path)
    return path


# ── Tests: load_excel_config ─────────────────────────────────────────

class TestLoadExcelConfig:
    _URL_HEADER = ["num", "lang", "url"]
    _KW_HEADER  = ["num", "lang", "text", "button_name"]

    def test_returns_url_and_keyword_lists(self, tmp_path):
        path = _make_two_sheet_excel(
            tmp_path,
            [self._URL_HEADER, [1, "English", "https://example.com"]],
            [self._KW_HEADER,  [1, "English", "gtm_click", "btn-a"]],
        )
        urls, kws = load_excel_config(path)
        assert len(urls) == 1
        assert len(kws)  == 1

    def test_url_fields(self, tmp_path):
        path = _make_two_sheet_excel(
            tmp_path,
            [self._URL_HEADER, [3, "English", "https://example.com"]],
            [self._KW_HEADER],
        )
        urls, _ = load_excel_config(path)
        assert urls[0].url  == "https://example.com"
        assert urls[0].lang == "en"
        assert urls[0].num  == 3        # num preserved as-is from Excel

    def test_url_num_preserved_not_renumbered(self, tmp_path):
        """All URLs with num=1 in Excel must stay num=1 — no auto-increment."""
        path = _make_two_sheet_excel(
            tmp_path,
            [
                self._URL_HEADER,
                [1, "English", "https://a.com"],
                [1, "English", "https://b.com"],
                [1, "English", "https://c.com"],
            ],
            [self._KW_HEADER],
        )
        urls, _ = load_excel_config(path)
        assert [u.num for u in urls] == [1, 1, 1]

    def test_keyword_fields(self, tmp_path):
        path = _make_two_sheet_excel(
            tmp_path,
            [self._URL_HEADER],
            [self._KW_HEADER, [2, "Traditional Chinese", "collect", "btn-b"]],
        )
        _, kws = load_excel_config(path)
        assert kws[0].text        == "collect"
        assert kws[0].lang        == "tc"
        assert kws[0].num         == 2
        assert kws[0].button_name == "btn-b"

    def test_keyword_missing_button_name(self, tmp_path):
        path = _make_two_sheet_excel(
            tmp_path,
            [self._URL_HEADER],
            [self._KW_HEADER, [1, "English", "pageview", None]],
        )
        _, kws = load_excel_config(path)
        assert kws[0].button_name is None

    def test_blank_url_rows_skipped(self, tmp_path):
        path = _make_two_sheet_excel(
            tmp_path,
            [self._URL_HEADER, [1, "English", None], [2, "English", "https://b.com"]],
            [self._KW_HEADER],
        )
        urls, _ = load_excel_config(path)
        assert len(urls) == 1
        assert urls[0].url == "https://b.com"

    def test_blank_keyword_rows_skipped(self, tmp_path):
        path = _make_two_sheet_excel(
            tmp_path,
            [self._URL_HEADER],
            [self._KW_HEADER, [1, "English", None], [2, "English", "collect", None]],
        )
        _, kws = load_excel_config(path)
        assert len(kws) == 1
        assert kws[0].text == "collect"

    def test_missing_sheet_raises_value_error(self, tmp_path):
        wb = openpyxl.Workbook()
        wb.active.title = "URLs"
        path = str(tmp_path / "bad.xlsx")
        wb.save(path)
        with pytest.raises(ValueError, match="header row|required column"):
            load_excel_config(path)

    def test_nonexistent_file_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            load_excel_config(str(tmp_path / "ghost.xlsx"))

    def test_language_aliases_are_normalized(self, tmp_path):
        path = _make_two_sheet_excel(
            tmp_path,
            [
                self._URL_HEADER,
                [1, "zh-hk", "https://tc.example.com"],
                [2, "schi", "https://sc.example.com"],
                [3, "eng", "https://en.example.com"],
            ],
            [
                self._KW_HEADER,
                [1, "chi", "kw_tc", "btn_tc"],
                [2, "zh-cn", "kw_sc", "btn_sc"],
                [3, "en-hk", "kw_en", "btn_en"],
            ],
        )
        urls, kws = load_excel_config(path)
        assert [u.lang for u in urls] == ["tc", "sc", "en"]
        assert [k.lang for k in kws] == ["tc", "sc", "en"]

    def test_more_language_aliases_are_normalized(self, tmp_path):
        path = _make_two_sheet_excel(
            tmp_path,
            [
                self._URL_HEADER,
                [1, "zh_tw", "https://tc.example.com"],
                [2, "zh-hans", "https://sc.example.com"],
                [3, "en_us", "https://en.example.com"],
            ],
            [
                self._KW_HEADER,
                [1, "traditional-chinese", "kw_tc", "btn_tc"],
                [2, "simplified_chinese", "kw_sc", "btn_sc"],
                [3, "EN-GB", "kw_en", "btn_en"],
            ],
        )
        urls, kws = load_excel_config(path)
        assert [u.lang for u in urls] == ["tc", "sc", "en"]
        assert [k.lang for k in kws] == ["tc", "sc", "en"]

    def test_new_template_auto_fallback_parses_to_models(self, tmp_path):
        path = _make_new_template_excel(
            tmp_path,
            [
                {
                    "B": "Required",
                    "C": "Required",
                    "E": "Required",
                    "F": "Required",
                    "H": "page",
                    "K": "/zh-cn",
                    "L": "/campaign/a",
                    "Q": "dc_type_a",
                    "R": "dc_main_a",
                    "S": "16998351256/labelA",
                    "V": "ttd_acc_a",
                    "W": "ttd_ct_a",
                    "X": "tab_acc_a",
                    "Y": "tab_en_a",
                },
                {
                    "D": "Required",
                    "G": "Required",
                    "H": "event",
                    "K": "/en-hk",
                    "L": "/campaign/b",
                    "P": "btn-apply",
                    "T": "meta_pixel_b",
                    "U": "meta_ev_b",
                    "AA": "applier_action_b",
                    "AB": "applier_track_b",
                },
            ],
        )
        urls, kws = load_excel_config(path)

        assert [(u.num, u.lang, u.url) for u in urls] == [
            (1, "sc", "https://www.hangseng.com/zh-cn/campaign/a"),
            (2, "en", "https://www.hangseng.com/en-hk/campaign/b"),
        ]
        assert (1, "sc", "dc_main_a", "dc_type_a", None, "dc") in [
            (k.num, k.lang, k.text, k.secondary_text, k.button_name, k.tag_type)
            for k in kws
        ]
        assert (1, "sc", "labelA", "16998351256", None, "gtag") in [
            (k.num, k.lang, k.text, k.secondary_text, k.button_name, k.tag_type)
            for k in kws
        ]
        assert (2, "en", "meta_ev_b", "meta_pixel_b", "btn-apply", "meta") in [
            (k.num, k.lang, k.text, k.secondary_text, k.button_name, k.tag_type)
            for k in kws
        ]
        assert (1, "sc", "ttd_ct_a", "ttd_acc_a", None, "ttd") in [
            (k.num, k.lang, k.text, k.secondary_text, k.button_name, k.tag_type)
            for k in kws
        ]
        assert (1, "sc", "tab_en_a", "tab_acc_a", None, "taboola") in [
            (k.num, k.lang, k.text, k.secondary_text, k.button_name, k.tag_type)
            for k in kws
        ]
        assert (2, "en", "applier_track_b", "applier_action_b", None, "btn-apply", "applier") in [
            (k.num, k.lang, k.text, k.secondary_text, k.tertiary_text, k.button_name, k.tag_type)
            for k in kws
        ]

    def test_new_template_duplicate_composite_key_is_allowed(self, tmp_path):
        path = _make_new_template_excel(
            tmp_path,
            [
                {
                    "B": "Required",
                    "H": "event",
                    "K": "/zh-hk",
                    "L": "/dup/path",
                    "P": "btn-1",
                    "Q": "dc_type_1",
                    "R": "dc_kw_1",
                },
                {
                    "C": "Required",
                    "H": "event",
                    "K": "zh_hk",
                    "L": "/dup/path",
                    "P": "btn-1",
                    "S": "16998350000/gtag_kw_2",
                },
            ],
        )
        urls, kws = load_excel_config(path)
        assert len(urls) == 1
        assert urls[0].num == 1
        assert urls[0].lang == "tc"
        assert urls[0].url.endswith("/zh-hk/dup/path")
        assert (1, "tc", "dc_kw_1", "dc_type_1", "btn-1", "dc") in [
            (k.num, k.lang, k.text, k.secondary_text, k.button_name, k.tag_type)
            for k in kws
        ]
        assert (1, "tc", "gtag_kw_2", "16998350000", "btn-1", "gtag") in [
            (k.num, k.lang, k.text, k.secondary_text, k.button_name, k.tag_type)
            for k in kws
        ]

    def test_new_template_cms_url_is_built_with_cms_lang_and_index(self, tmp_path):
        path = _make_new_template_excel(
            tmp_path,
            [
                {
                    "C": "Required",
                    "H": "page",
                    "K": "/zh-hk",
                    "L": "/cms/personal/credit-cards",
                    "S": "16998351256/label_cms",
                },
            ],
        )
        urls, kws = load_excel_config(path)

        assert [(u.num, u.lang, u.url) for u in urls] == [
            (1, "tc", "https://cms.hangseng.com/cms/personal/credit-cards/chi/index.html"),
        ]
        assert (1, "tc", "label_cms", "16998351256", None, "gtag") in [
            (k.num, k.lang, k.text, k.secondary_text, k.button_name, k.tag_type)
            for k in kws
        ]

    def test_new_template_all_language_expands_to_three_urls_and_keywords(self, tmp_path):
        path = _make_new_template_excel(
            tmp_path,
            [
                {
                    "C": "Required",
                    "E": "Required",
                    "F": "Required",
                    "G": "Required",
                    "H": "page",
                    "K": "(All)",
                    "L": "/promo/all",
                    "S": "16998351256/label_all",
                    "V": "ttd_acc_all",
                    "W": "ttd_ct_all",
                    "X": "tab_acc_all",
                    "Y": "tab_en_all",
                    "AA": "applier_action_all",
                    "AB": "applier_track_all",
                },
            ],
        )
        urls, kws = load_excel_config(path)

        assert [(u.lang, u.url) for u in urls] == [
            ("tc", "https://www.hangseng.com/zh-hk/promo/all"),
            ("sc", "https://www.hangseng.com/zh-cn/promo/all"),
            ("en", "https://www.hangseng.com/en-hk/promo/all"),
        ]
        assert len(kws) == 12
        assert {(k.lang, k.text, k.secondary_text) for k in kws if k.tag_vendor == "gtag"} == {
            ("tc", "label_all", "16998351256"),
            ("sc", "label_all", "16998351256"),
            ("en", "label_all", "16998351256"),
        }
        assert {(k.lang, k.text, k.secondary_text) for k in kws if k.tag_vendor == "ttd"} == {
            ("tc", "ttd_ct_all", "ttd_acc_all"),
            ("sc", "ttd_ct_all", "ttd_acc_all"),
            ("en", "ttd_ct_all", "ttd_acc_all"),
        }
        assert {(k.lang, k.text, k.secondary_text) for k in kws if k.tag_vendor == "taboola"} == {
            ("tc", "tab_en_all", "tab_acc_all"),
            ("sc", "tab_en_all", "tab_acc_all"),
            ("en", "tab_en_all", "tab_acc_all"),
        }
        assert {(k.lang, k.text, k.secondary_text, k.tertiary_text) for k in kws if k.tag_vendor == "applier"} == {
            ("tc", "applier_track_all", "applier_action_all", None),
            ("sc", "applier_track_all", "applier_action_all", None),
            ("en", "applier_track_all", "applier_action_all", None),
        }


# ── Tests: convert_excel_to_json_data ────────────────────────────────

class TestConvertExcelToJsonData:
    def _make_excel(self, tmp_path, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in rows:
            ws.append(row)
        path = str(tmp_path / "kw.xlsx")
        wb.save(path)
        return path

    def test_basic_keywords(self, tmp_path):
        path = self._make_excel(
            tmp_path, [["Keyword"], ["collect"], ["gtm.js"]]
        )
        assert convert_excel_to_json_data(path) == ["collect", "gtm.js"]

    def test_empty_rows_skipped(self, tmp_path):
        path = self._make_excel(
            tmp_path, [["Keyword"], ["collect"], [None], ["gtm.js"]]
        )
        assert convert_excel_to_json_data(path) == ["collect", "gtm.js"]

    def test_only_header_returns_empty(self, tmp_path):
        path = self._make_excel(tmp_path, [["Keyword"]])
        assert convert_excel_to_json_data(path) == []

    def test_values_are_stripped(self, tmp_path):
        path = self._make_excel(tmp_path, [["Keyword"], ["  spaces  "]])
        assert convert_excel_to_json_data(path) == ["spaces"]
