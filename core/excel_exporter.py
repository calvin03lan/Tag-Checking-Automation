"""
Excel report exporter.

This module is intentionally isolated from UI and automation code.
It only receives result rows (including screenshot paths) and writes
an Excel file in the expected report format.
"""
import os
import shutil
from datetime import datetime
from typing import Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models.config import LANG_MAP_INV
from models.session import ReportEntry

# ── Colour palette ──────────────────────────────────────────────────
_PASS_BG = "FF92D050"      # green
_FAILED_BG = "FFFF0000"    # red
_HEADER_BG = "FF4472C4"    # blue
_WHITE = "FFFFFFFF"

# ── Column layout ───────────────────────────────────────────────────
# Columns: URL # | URL | URL Lang | KW # | Keyword | KW Lang | ID | Result | Tested At | Screenshot
_HEADERS = [
    "URL #", "URL", "URL Lang",
    "KW #", "Keyword", "KW Lang", "ID",
    "Result", "Tested At", "Screenshot",
]
_COL_WIDTHS = {1: 8, 2: 55, 3: 18, 4: 8, 5: 30, 6: 12, 7: 20, 8: 12, 9: 22}
# col 10 (Screenshot) width set dynamically when an image is embedded

# Keep existing visual row height behavior for rows with screenshots.
_IMG_TARGET_H = 180
_IMG_ROW_H = 145
_DEFAULT_ROW_H = 30
_SCREENSHOT_COL = 10
_RESULT_COL = 8

_VENDOR_ORDER = ["doubleclick", "gtag", "meta", "ttd", "taboola", "applier"]
_VENDOR_LABELS = {
    "doubleclick": "DoubleClick",
    "gtag": "Gtag",
    "meta": "Meta",
    "ttd": "The Trade Desk",
    "taboola": "Taboola",
    "applier": "Applier",
}
_VENDOR_TO_FIRST_SHEET_COL = {
    "doubleclick": 2,  # B
    "gtag": 3,         # C
    "meta": 4,         # D
    "ttd": 5,          # E
    "taboola": 6,      # F
    "applier": 7,      # G
}
_DATA_START_ROW = 4
_LANG_COL = 11                 # K
_URL_PATH_COL = 12             # L
_FIRST_SHEET_HDR_ROW = 3
_FIRST_SHEET_START_COL = 29    # AC
_FIRST_SHEET_ROW_H = 90
_FIRST_SHEET_IMG_COL_W = 51
_FIRST_SHEET_LANG_ORDER = ["tc", "sc", "en"]


def _thin_border() -> Border:
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _pixels_to_excel_width(px: float) -> float:
    """
    Approximate pixel width to Excel column width units.
    """
    if px <= 12:
        return px / 12.0
    return (px - 5.0) / 7.0


def _excel_width_to_pixels(width: float) -> int:
    if width <= 0:
        return 1
    if width < 1:
        return max(1, int(round(width * 12.0)))
    return max(1, int(round(width * 7.0 + 5.0)))


def _points_to_pixels(points: float) -> int:
    if points <= 0:
        return 1
    return max(1, int(round(points * 96.0 / 72.0)))


class ExcelReportExporter:
    """Export report rows to a formatted Excel file."""

    def export_into_input_copy(
        self,
        entries: List[ReportEntry],
        input_excel_path: str,
        output_path: str,
    ) -> str:
        shutil.copy2(input_excel_path, output_path)
        wb = load_workbook(output_path)
        try:
            ws_test_result = self._prepare_test_result_sheet(wb)
            self._write_results_back_to_first_sheet(ws_test_result, entries)
            self._write_evidence_back_to_first_sheet(ws_test_result, entries)
            if "Evidence" in wb.sheetnames:
                del wb["Evidence"]
            self._prepare_summary_sheet(wb, entries)
            self._move_sheet_to_front(wb, "Test Result")
            self._move_sheet_to_front(wb, "Summary")
            wb.active = 0
            wb.save(output_path)
        finally:
            wb.close()
        return os.path.abspath(output_path)

    def _prepare_test_result_sheet(self, wb: Workbook):
        if "Test Result" in wb.sheetnames:
            del wb["Test Result"]
        source_ws = wb.worksheets[0]
        result_ws = wb.copy_worksheet(source_ws)
        result_ws.title = "Test Result"
        return result_ws

    def _prepare_summary_sheet(self, wb: Workbook, entries: List[ReportEntry]):
        if "Summary" in wb.sheetnames:
            del wb["Summary"]
        ws = wb.create_sheet("Summary")
        self._write_summary_overview(ws, entries)
        return ws

    def _write_summary_overview(self, ws, entries: List[ReportEntry]) -> None:
        ws.cell(row=1, column=1, value="Summary")
        ws.cell(row=2, column=1, value="Latest Update")
        ws.cell(row=3, column=1, value="Status")
        ws.cell(row=2, column=2, value=datetime.now().strftime("%Y/%m/%d"))

        if entries and all((entry.result or "").upper() == "PASS" for entry in entries):
            status_cell = ws.cell(row=3, column=2)
            try:
                status_cell.value = CellRichText(
                    "All ",
                    TextBlock(InlineFont(b=True, color="FF00B050"), "PASS"),
                )
            except Exception:
                status_cell.value = "All PASS"
                status_cell.font = Font(bold=True, color="FF00B050")

    def _move_sheet_to_front(self, wb: Workbook, name: str) -> None:
        if name not in wb.sheetnames:
            return
        idx = wb.sheetnames.index(name)
        ws = wb.worksheets[idx]
        wb._sheets.insert(0, wb._sheets.pop(wb._sheets.index(ws)))

    def export(self, entries: List[ReportEntry], output_path: str) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tag QA Report"

        self._write_header(ws)
        for row_idx, entry in enumerate(entries, start=2):
            self._write_entry(ws, row_idx, entry)

        self._set_column_widths(ws)
        self._write_new_template_sheet(wb, entries)
        wb.save(output_path)
        return os.path.abspath(output_path)

    def _write_new_template_sheet(
        self,
        wb: Workbook,
        entries: List[ReportEntry],
        sheet_name: str = "New Template Report",
    ) -> None:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)
        self._write_new_template_header(ws)
        grouped = self._group_entries_for_new_template(entries)
        for row_idx, row_key in enumerate(sorted(grouped.keys()), start=2):
            self._write_new_template_row(ws, row_idx, grouped[row_key])

    def _write_results_back_to_first_sheet(self, ws, entries: List[ReportEntry]) -> None:
        by_row_vendor = self._group_entries_for_new_template(entries)
        for source_row, vendor_map in by_row_vendor.items():
            if source_row <= 0:
                continue
            for vendor, entry in vendor_map.items():
                col = _VENDOR_TO_FIRST_SHEET_COL.get(vendor)
                if not col:
                    continue
                cell = ws.cell(row=source_row, column=col)
                cell.value = entry.result
                if entry.result == "PASS":
                    cell.fill = PatternFill("solid", fgColor=_PASS_BG)
                    cell.font = Font(bold=True)
                elif entry.result == "FAILED":
                    cell.fill = PatternFill("solid", fgColor=_FAILED_BG)
                    cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    def _write_evidence_back_to_first_sheet(self, ws, entries: List[ReportEntry]) -> None:
        grouped = self._group_entries_for_first_sheet_evidence(entries)
        row_map = self._expand_rows_for_all_language(ws, grouped.keys())
        self._write_first_sheet_evidence_header(ws)

        for source_row in sorted(grouped.keys()):
            target_rows = row_map.get(source_row)
            if not target_rows:
                continue
            self._write_first_sheet_evidence_rows(
                ws,
                target_rows=target_rows,
                lang_vendor_map=grouped[source_row],
            )

    def _group_entries_for_first_sheet_evidence(
        self,
        entries: List[ReportEntry],
    ) -> Dict[int, Dict[str, Dict[str, ReportEntry]]]:
        grouped: Dict[int, Dict[str, Dict[str, ReportEntry]]] = {}
        for entry in entries:
            if entry.source_row <= 0:
                continue
            vendor = entry.tag_vendor or "other"
            if vendor not in _VENDOR_ORDER:
                continue
            lang = (entry.url_lang or "").strip().lower()
            if not lang:
                continue

            by_lang = grouped.setdefault(entry.source_row, {})
            vendor_map = by_lang.setdefault(lang, {})
            existing = vendor_map.get(vendor)
            if existing is None or (existing.result == "PASS" and entry.result == "FAILED"):
                vendor_map[vendor] = entry
        return grouped

    def _expand_rows_for_all_language(self, ws, source_rows) -> Dict[int, List[int]]:
        row_map: Dict[int, List[int]] = {}
        offset = 0
        for source_row in sorted(r for r in source_rows if r > 0):
            current_row = source_row + offset
            if self._row_language_is_all(ws, current_row):
                ws.insert_rows(current_row + 1, amount=2)
                row_map[source_row] = [current_row, current_row + 1, current_row + 2]
                offset += 2
            else:
                row_map[source_row] = [current_row]
        return row_map

    def _row_language_is_all(self, ws, row: int) -> bool:
        raw = ws.cell(row=row, column=_LANG_COL).value
        if raw is None:
            return False
        normalized = str(raw).strip().lower()
        return normalized in {"all", "(all)"}

    def _write_first_sheet_evidence_header(self, ws) -> None:
        headers: List[str] = []
        for vendor in _VENDOR_ORDER:
            headers.extend([_VENDOR_LABELS[vendor], "Evidence"])

        for idx, text in enumerate(headers):
            col = _FIRST_SHEET_START_COL + idx
            cell = ws.cell(row=_FIRST_SHEET_HDR_ROW, column=col, value=text)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="bottom")
            if idx % 2 == 1:
                ws.column_dimensions[get_column_letter(col)].width = _FIRST_SHEET_IMG_COL_W

    def _write_first_sheet_evidence_rows(
        self,
        ws,
        target_rows: List[int],
        lang_vendor_map: Dict[str, Dict[str, ReportEntry]],
    ) -> None:
        if not target_rows:
            return

        if len(target_rows) == 1:
            lang = self._pick_single_language(lang_vendor_map)
            vendor_map = lang_vendor_map.get(lang, {})
            self._write_first_sheet_evidence_row(ws, target_rows[0], vendor_map)
            return

        for idx, lang in enumerate(_FIRST_SHEET_LANG_ORDER):
            row = target_rows[idx] if idx < len(target_rows) else target_rows[-1]
            vendor_map = lang_vendor_map.get(lang, {})
            self._write_first_sheet_evidence_row(ws, row, vendor_map)

    def _pick_single_language(self, lang_vendor_map: Dict[str, Dict[str, ReportEntry]]) -> str:
        for lang in _FIRST_SHEET_LANG_ORDER:
            if lang in lang_vendor_map:
                return lang
        if not lang_vendor_map:
            return ""
        return sorted(lang_vendor_map.keys())[0]

    def _write_first_sheet_evidence_row(
        self,
        ws,
        row: int,
        vendor_map: Dict[str, ReportEntry],
    ) -> None:
        ws.row_dimensions[row].height = _FIRST_SHEET_ROW_H

        for idx, vendor in enumerate(_VENDOR_ORDER):
            result_col = _FIRST_SHEET_START_COL + idx * 2
            evidence_col = result_col + 1

            result_cell = ws.cell(row=row, column=result_col)
            result_cell.alignment = Alignment(horizontal="center", vertical="center")
            evidence_cell = ws.cell(row=row, column=evidence_col)
            evidence_cell.alignment = Alignment(vertical="center")

            entry = vendor_map.get(vendor)
            if entry is None:
                continue

            result_cell.value = entry.result
            if entry.result == "PASS":
                result_cell.fill = PatternFill("solid", fgColor=_PASS_BG)
                result_cell.font = Font(bold=True)
            elif entry.result == "FAILED":
                result_cell.fill = PatternFill("solid", fgColor=_FAILED_BG)
                result_cell.font = Font(bold=True)
            if not entry.screenshot_path or not os.path.exists(entry.screenshot_path):
                continue

            try:
                img = XlImage(entry.screenshot_path)
                col_letter = get_column_letter(evidence_col)
                col_width = ws.column_dimensions[col_letter].width or _FIRST_SHEET_IMG_COL_W
                row_height = ws.row_dimensions[row].height or _FIRST_SHEET_ROW_H
                img.width = _excel_width_to_pixels(float(col_width))
                img.height = _points_to_pixels(float(row_height))
                ws.add_image(img, f"{get_column_letter(evidence_col)}{row}")
            except Exception:
                evidence_cell.value = "[screenshot error]"

    def _write_new_template_header(self, ws) -> None:
        hdr_fill = PatternFill("solid", fgColor=_HEADER_BG)
        hdr_font = Font(bold=True, color=_WHITE)
        headers: List[str] = []
        for vendor in _VENDOR_ORDER:
            label = _VENDOR_LABELS[vendor]
            headers.extend([label, "Evidence"])

        for col, text in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=text)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin_border()
        ws.row_dimensions[1].height = 22

        for idx in range(len(_VENDOR_ORDER)):
            result_col = 1 + idx * 2
            evidence_col = result_col + 1
            ws.column_dimensions[get_column_letter(result_col)].width = _COL_WIDTHS[_RESULT_COL]
            ws.column_dimensions[get_column_letter(evidence_col)].width = 72.50

    def _group_entries_for_new_template(
        self,
        entries: List[ReportEntry],
    ) -> Dict[int, Dict[str, ReportEntry]]:
        grouped: Dict[int, Dict[str, ReportEntry]] = {}
        for idx, entry in enumerate(entries, start=1):
            row_key = entry.source_row if entry.source_row > 0 else 1_000_000 + idx
            vendor = entry.tag_vendor or "other"
            if vendor not in _VENDOR_ORDER:
                continue
            if row_key not in grouped:
                grouped[row_key] = {}

            existing = grouped[row_key].get(vendor)
            if existing is None or (existing.result == "PASS" and entry.result == "FAILED"):
                grouped[row_key][vendor] = entry
        return grouped

    def _write_new_template_row(self, ws, row: int, vendor_map: Dict[str, ReportEntry]) -> None:
        row_has_image = False
        for idx, vendor in enumerate(_VENDOR_ORDER):
            result_col = 1 + idx * 2
            evidence_col = result_col + 1
            entry = vendor_map.get(vendor)

            result_value = entry.result if entry else "N/A"
            result_cell = ws.cell(row=row, column=result_col, value=result_value)
            result_cell.alignment = Alignment(horizontal="center", vertical="center")
            result_cell.border = _thin_border()
            if result_value in {"PASS", "FAILED"}:
                result_cell.fill = PatternFill(
                    "solid", fgColor=_PASS_BG if result_value == "PASS" else _FAILED_BG
                )
                result_cell.font = Font(bold=True)

            evidence_cell = ws.cell(row=row, column=evidence_col)
            evidence_cell.alignment = Alignment(vertical="center")
            evidence_cell.border = _thin_border()

            if not entry or not entry.screenshot_path or not os.path.exists(entry.screenshot_path):
                evidence_cell.value = "N/A"
                continue

            try:
                img = XlImage(entry.screenshot_path)
                if img.height > 0:
                    scale = _IMG_TARGET_H / float(img.height)
                    img.height = int(_IMG_TARGET_H)
                    img.width = int(img.width * scale)
                anchor = f"{get_column_letter(evidence_col)}{row}"
                ws.add_image(img, anchor)
                row_has_image = True
            except Exception:
                evidence_cell.value = "[screenshot error]"

        ws.row_dimensions[row].height = _IMG_ROW_H if row_has_image else _DEFAULT_ROW_H

    def _write_header(self, ws) -> None:
        hdr_fill = PatternFill("solid", fgColor=_HEADER_BG)
        hdr_font = Font(bold=True, color=_WHITE)
        for col, text in enumerate(_HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=text)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin_border()
        ws.row_dimensions[1].height = 22

    def _write_entry(self, ws, row: int, entry: ReportEntry) -> None:
        is_pass = entry.result == "PASS"
        res_fill = PatternFill("solid", fgColor=_PASS_BG if is_pass else _FAILED_BG)

        url_lang_label = LANG_MAP_INV.get(entry.url_lang, entry.url_lang.upper())
        kw_lang_label = LANG_MAP_INV.get(entry.kw_lang, entry.kw_lang.upper())

        values = [
            entry.url_index,          # col 1  URL #
            entry.url,                # col 2  URL
            url_lang_label,           # col 3  URL Lang
            entry.kw_num,             # col 4  KW #
            entry.kw_text,            # col 5  Keyword
            kw_lang_label,            # col 6  KW Lang
            entry.kw_button or "—",   # col 7  ID
            entry.result,             # col 8  Result
            entry.tested_at,          # col 9  Tested At
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = _thin_border()
            if col == _RESULT_COL:
                cell.fill = res_fill
                cell.font = Font(bold=True)

        has_shot = bool(
            entry.screenshot_path and os.path.exists(entry.screenshot_path)
        )
        if has_shot:
            try:
                img = XlImage(entry.screenshot_path)
                # Preserve aspect ratio while keeping screenshot rows at
                # the existing row-height configuration.
                if img.height > 0:
                    scale = _IMG_TARGET_H / float(img.height)
                    img.height = int(_IMG_TARGET_H)
                    img.width = int(img.width * scale)
                ws.add_image(img, f"J{row}")
                ws.row_dimensions[row].height = _IMG_ROW_H
                new_col_w = _pixels_to_excel_width(img.width) + 1.2
                curr_col_w = ws.column_dimensions["J"].width
                ws.column_dimensions["J"].width = max(curr_col_w or 0, new_col_w)
            except Exception:
                ws.cell(row=row, column=_SCREENSHOT_COL, value="[screenshot error]")
                ws.row_dimensions[row].height = _DEFAULT_ROW_H
        else:
            ws.cell(row=row, column=_SCREENSHOT_COL, value="N/A")
            ws.row_dimensions[row].height = _DEFAULT_ROW_H

    def _set_column_widths(self, ws) -> None:
        for col, width in _COL_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col)].width = width
